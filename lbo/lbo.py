import pandas as pd
import numpy as np
from openpyxl import Workbook
import openpyxl as opx
from openpyxl.utils.dataframe import dataframe_to_rows
from utility.dateUtils import get_month_list
from database.dbPrices import get_historical_lmp

from dateutil.relativedelta import relativedelta
from datetime import date, datetime
import sys




LBO_FSLI_COLOR = {'Delivered Fuel Expense':-1,'Variable O&M Expense':-1,'Net Emissions Expense':-1}


LBO_SUM_FSLIS = ['Energy Revenue','Delivered Fuel Expense','Variable O&M Expense',
                 'Net Emissions Expense','Gross Energy Margin','Hedges',
                 'Net Energy Margin','Fixed Fuel Transport','Capacity Revenue',
                 'Ancillary Revenue','Other Revenue','Gross Margin',
                 'FOM','Taxes','Insurance','Fixed Costs',
                 'EBITDA','Capex','EBITDA less Capex',
                 'Generation', 'Generation - On Peak', 'Generation - Off Peak',
                 'Hours - On Peak', 'Hours - Off Peak']




def convert_date(datetimeobj):

    if isinstance(datetimeobj, date):
        return datetimeobj

    if isinstance(datetimeobj, datetime):
        return datetimeobj.date()

    print (datetimeobj)
    sys.exit()



def build_lbo_financials(powerplant_df, portfolio, scenario, version, dispatch_df, lbo_assumptions_df):
    lbo_financials_df = pd.DataFrame()

    dispatch_financials_fsli_list = ['Generation - On Peak',
                                     'Generation - Off Peak',
                                     'Generation',
                                     'ICAP',
                                     'Capacity Factor',
                                     'Capacity Factor - On Peak',
                                     'Capacity Factor - Off Peak',
                                     'Realized Power Price - Off Peak',
                                     'Realized Power Price - On Peak',
                                     'Realized Fuel Price - Off Peak',
                                     'Realized Fuel Price - On Peak',
                                     'Realized Spread - ATC',
                                     'Realized Spread - Off Peak',
                                     'Realized Spread - On Peak',
                                     'Energy Revenue',
                                     'Delivered Fuel Expense',
                                     'Variable O&M Expense',
                                     'Net Emissions Expense',
                                     'on_hours',
                                     'off_hours']

    lbo_dispatch_df = dispatch_df.loc[dispatch_df.fsli.isin(dispatch_financials_fsli_list)]

    lbo_financials_df = lbo_dispatch_df

    dispatch_plant_list = list(set(list(lbo_financials_df.entity)))

    """ company, scenario, version, entity, fsli, period, value """

    for fsli in ['Capacity Revenue','FOM','Taxes','Insurance','Fixed Costs','Hedges','Fixed Fuel Transport','Other Revenue','Ancillary Revenue','Capex']:
        for index, row in powerplant_df.iterrows():
            entity = row['name']

            if fsli in ['ICAP']:
                if entity not in dispatch_plant_list:
                    lbo_fsli_entity_assumptions_df = lbo_assumptions_df.loc[(lbo_assumptions_df.entity == entity) & (lbo_assumptions_df.fsli == fsli)]
                    unit = lbo_fsli_entity_assumptions_df.iloc[0]['unit']
                    temp_fsli_df = pd.DataFrame()
                    if unit == '$':
                        temp_fsli_df = lbo_fsli_entity_assumptions_df[['entity', 'fsli', 'period', 'value']]
                        temp_fsli_df['company'] = portfolio
                        temp_fsli_df['scenario'] = scenario
                        temp_fsli_df['version'] = version
                    lbo_financials_df = lbo_financials_df.append(temp_fsli_df)
                continue

            lbo_fsli_entity_assumptions_df = lbo_assumptions_df.loc[(lbo_assumptions_df.entity == entity) & (lbo_assumptions_df.fsli == fsli)]
            unit = lbo_fsli_entity_assumptions_df.iloc[0]['unit']
            temp_fsli_df = pd.DataFrame()
            if unit == '$':
                temp_fsli_df = lbo_fsli_entity_assumptions_df[['entity', 'fsli', 'period', 'value']]
                temp_fsli_df['company'] = portfolio
                temp_fsli_df['scenario'] = scenario
                temp_fsli_df['version'] = version
            lbo_financials_df = lbo_financials_df.append(temp_fsli_df)

    lbo_financials_df['scenario'] = scenario
    lbo_financials_df['version'] = version

    lbo_financials_df['period'] = lbo_financials_df.apply(lambda row: convert_date(row['period']), axis=1)

    lbo_financials_df = lbo_financials_df[['company','scenario','version','entity','fsli','period','value']]

    pivot_lbo_financials_df = pd.pivot_table(lbo_financials_df, index=['company','scenario','version','entity', 'period'], columns=['fsli'], values='value', aggfunc=np.sum)

    # lbo_financials_df.to_csv("lbo_financials_df.csv")

    pivot_lbo_financials_df = pivot_lbo_financials_df.reset_index()

    pivot_lbo_financials_df.fillna(0.0, inplace=True)

    pivot_lbo_financials_df['Gross Energy Margin'] = pivot_lbo_financials_df['Energy Revenue'] - pivot_lbo_financials_df['Delivered Fuel Expense'] - pivot_lbo_financials_df['Variable O&M Expense'] - pivot_lbo_financials_df['Net Emissions Expense']

    pivot_lbo_financials_df['Net Energy Margin'] = pivot_lbo_financials_df['Gross Energy Margin'] + pivot_lbo_financials_df['Hedges']

    pivot_lbo_financials_df['Gross Margin'] = pivot_lbo_financials_df['Net Energy Margin'] + \
                                              pivot_lbo_financials_df['Fixed Fuel Transport'] + \
                                              pivot_lbo_financials_df['Capacity Revenue'] + \
                                              pivot_lbo_financials_df['Ancillary Revenue'] + \
                                              pivot_lbo_financials_df['Other Revenue']

    pivot_lbo_financials_df['EBITDA'] = pivot_lbo_financials_df['Gross Margin'] + \
                                                pivot_lbo_financials_df['Fixed Costs']

    pivot_lbo_financials_df['EBITDA less Capex'] = pivot_lbo_financials_df['EBITDA'] + \
                                                        pivot_lbo_financials_df['Capex']


    pivot_lbo_financials_df['Realized Power Price'] = pivot_lbo_financials_df['Energy Revenue'] / \
                                                        pivot_lbo_financials_df['Generation']


    pivot_lbo_financials_df.rename(columns={'on_hours':'Hours - On Peak', 'off_hours':'Hours - Off Peak'}, inplace=True)


    pivot_lbo_financials_df = pivot_lbo_financials_df.reset_index()

    pivot_lbo_financials_df = pivot_lbo_financials_df[[item for item in pivot_lbo_financials_df.columns if item != 'index']]

    retirement_date_df = powerplant_df[['name','retirement_date']]

    pivot_lbo_financials_df = pd.merge(pivot_lbo_financials_df, retirement_date_df, left_on=['entity'], right_on=['name'], how='left')

    """ if a plant is retired, just drop that row """

    for index, row in pivot_lbo_financials_df.iterrows():
        if row['period'] > row['retirement_date']:
            pivot_lbo_financials_df.drop(index, inplace=True)

    # pivot_lbo_financials_df.to_csv("pivot_lbo_financials_df.csv")



    melted_pivot_lbo_financials_df = pd.melt(pivot_lbo_financials_df, id_vars=['company','scenario','version','entity','period'],
                                             value_vars=[item for item in list(pivot_lbo_financials_df.columns) if item not in ['company','scenario','version','entity','period']],
                                             var_name='fsli',
                                             value_name='value')


    melted_pivot_lbo_financials_df.rename(columns={'company':'portfolio'}, inplace=True)

    # melted_pivot_lbo_financials_df.to_csv("melted_pivot_lbo_financials_df.csv")

    return melted_pivot_lbo_financials_df



def write_lbo_financials_report_monthly(dest_file_path, lbo_financials_df, portfolio):

    wb = Workbook()
    entity_list = list(sorted(list(set(list(lbo_financials_df['entity'])))))


    # step 1, apply lbo color for sinage
    pivot_lbo_financials_df = pd.pivot_table(lbo_financials_df, index=['portfolio','scenario','version','entity','period'], columns=['fsli'], values='value', aggfunc=np.sum)

    for fsli in LBO_FSLI_COLOR:
        pivot_lbo_financials_df[fsli] = pivot_lbo_financials_df[fsli] * LBO_FSLI_COLOR[fsli]

    pivot_lbo_financials_df = pivot_lbo_financials_df.reset_index()

    lbo_financials_df = pd.melt(pivot_lbo_financials_df, id_vars=['portfolio','scenario','version','entity','period'],
                                value_vars=[item for item in list(pivot_lbo_financials_df.columns) if item not in ['portfolio','scenario','version','entity','period']],
                                var_name='fsli',
                                value_name='value')

    lbo_financials_df = lbo_financials_df.reset_index()

    pnl_lbo_financials_df = lbo_financials_df.loc[lbo_financials_df.fsli.isin(['Energy Revenue',
                                                                                  'Delivered Fuel Expense',
                                                                                  'Variable O&M Expense',
                                                                                  'Net Emissions Expense',
                                                                                  'Gross Energy Margin',
                                                                                  'Hedges',
                                                                                  'Net Energy Margin',
                                                                                  'Fixed Fuel Transport',
                                                                                  'Capacity Revenue',
                                                                                  'Ancillary Revenue',
                                                                                  'Other Revenue',
                                                                                  'Gross Margin',
                                                                                  'FOM',
                                                                                  'Taxes',
                                                                                  'Insurance',
                                                                                  'Fixed Costs',
                                                                                  'EBITDA',
                                                                                  'Capex',
                                                                                  'EBITDA less Capex'])]

    summary_df = pd.pivot_table(pnl_lbo_financials_df, index=['portfolio','scenario','version','fsli'], columns=['period'], values='value', aggfunc=np.sum)
    summary_df = summary_df.reset_index()

    summary_df = summary_df[[column for column in summary_df.columns if column not in ['portfolio','scenario','version']]]

    summary_df.rename(columns={'fsli': portfolio}, inplace=True)

    summary_df = summary_df.set_index(portfolio)



    summary_df = summary_df.reindex(['Energy Revenue','Delivered Fuel Expense','Variable O&M Expense',
                                     'Net Emissions Expense','Gross Energy Margin','Hedges',
                                     'Net Energy Margin','Fixed Fuel Transport','Capacity Revenue',
                                     'Ancillary Revenue','Other Revenue','Gross Margin',
                                     'FOM','Taxes','Insurance','Fixed Costs',
                                     'EBITDA','Capex','EBITDA less Capex'])


    capacity_row_group_df = lbo_financials_df.loc[lbo_financials_df.fsli.isin(['ICAP', 'Generation', 'Generation - On Peak', 'Generation - Off Peak', 'Hours - On Peak', 'Hours - Off Peak'])]
    capacity_row_group_df = pd.pivot_table(capacity_row_group_df, index=['portfolio','scenario','version', 'period'], columns=['fsli'], values='value', aggfunc=np.sum)
    capacity_row_group_df = capacity_row_group_df.reset_index()

    capacity_row_group_df['Capacity Factor'] = capacity_row_group_df.apply(lambda row: row['Generation'] / (row['ICAP'] *  24 * row['period'].day), axis=1)
    capacity_row_group_df['Capacity Factor - On Peak'] = capacity_row_group_df['Generation - On Peak'] / (capacity_row_group_df['ICAP'] * capacity_row_group_df['Hours - On Peak'] / len(entity_list))
    capacity_row_group_df['Capacity Factor - Off Peak'] = capacity_row_group_df['Generation - Off Peak'] / (capacity_row_group_df['ICAP'] * capacity_row_group_df['Hours - Off Peak'] / len(entity_list))
    capacity_row_group_df['Hours - On Peak'] = capacity_row_group_df['Hours - On Peak'] / len(entity_list)
    capacity_row_group_df['Hours - Off Peak'] = capacity_row_group_df['Hours - Off Peak'] / len(entity_list)


    # capacity_row_group_df.rename(columns={'on_hours':'Hours - On Peak','off_hours':"Hours - Off Peak"}, inplace=True)


    capacity_row_group_df = pd.melt(capacity_row_group_df, id_vars=['portfolio','scenario','version','period'],
                                    value_vars=[item for item in list(capacity_row_group_df.columns) if item not in ['portfolio','scenario','version','period']],
                                    var_name='fsli',
                                    value_name='value')

    capacity_row_group_df = pd.pivot_table(capacity_row_group_df, index=['portfolio','scenario','version','fsli'], columns=['period'], values='value', aggfunc=np.sum )
    capacity_row_group_df = capacity_row_group_df.reset_index()


    capacity_row_group_df = capacity_row_group_df[[item for item in capacity_row_group_df.columns if item not in ['portfolio', 'scenario', 'version']]]
    capacity_row_group_df.rename(columns={'fsli':portfolio}, inplace=True)

    capacity_row_group_df = capacity_row_group_df.set_index(portfolio)

    capacity_row_group_df = capacity_row_group_df.reindex(['ICAP', 'Generation', 'Generation - On Peak', 'Generation - Off Peak', 'Hours - On Peak', 'Hours - Off Peak'])

    summary_df = summary_df.append(capacity_row_group_df)


    summary_df.rename(columns={'fsli': portfolio}, inplace=True)


    summary_df = summary_df.reset_index()
    summary_tab = wb.copy_worksheet(wb.active)
    summary_tab.title = 'Summary'
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        summary_tab.append(r)



    # annual consolidated view tab

    annual_consolidated_tab = wb.copy_worksheet(wb.active)

    sum_fslis_lbo_financials_df = lbo_financials_df.loc[lbo_financials_df.fsli.isin(LBO_SUM_FSLIS)]
    sum_fslis_lbo_financials_df.loc[:,'year'] = pd.DatetimeIndex(sum_fslis_lbo_financials_df['period']).year
    grouped_sum_fslis_lbo_financials_df = sum_fslis_lbo_financials_df.groupby(['portfolio','scenario','version','entity','fsli','year']).sum()
    grouped_sum_fslis_lbo_financials_df = grouped_sum_fslis_lbo_financials_df.reset_index()


    average_fslis = ['ICAP']
    average_fslis_lbo_financials_df = lbo_financials_df.loc[lbo_financials_df.fsli.isin(average_fslis)]
    average_fslis_lbo_financials_df.loc[:,'year'] = pd.DatetimeIndex(average_fslis_lbo_financials_df['period']).year
    grouped_average_fslis_lbo_financials_df = average_fslis_lbo_financials_df.groupby(['portfolio','scenario','version','entity','fsli','year']).mean()
    grouped_average_fslis_lbo_financials_df = grouped_average_fslis_lbo_financials_df.reset_index()
    annual_lbo_financials_df = grouped_sum_fslis_lbo_financials_df.append(grouped_average_fslis_lbo_financials_df)

    # annual_lbo_financials_df.to_csv("annual_lbo_financials_df.csv")

    pivot_annual_lbo_financials_df = pd.pivot_table(annual_lbo_financials_df, index=['portfolio','scenario','version','entity','year'], columns=['fsli'], values='value', aggfunc=np.sum)

    pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reset_index()

    pivot_annual_lbo_financials_df['Capacity Factor'] = pivot_annual_lbo_financials_df['Generation'] / ( pivot_annual_lbo_financials_df['ICAP'] * ( pivot_annual_lbo_financials_df['Hours - On Peak'] + pivot_annual_lbo_financials_df['Hours - Off Peak'] ))
    pivot_annual_lbo_financials_df['Capacity Factor - On Peak'] = pivot_annual_lbo_financials_df['Generation - On Peak'] / ( pivot_annual_lbo_financials_df['ICAP'] * ( pivot_annual_lbo_financials_df['Hours - On Peak'] ))
    pivot_annual_lbo_financials_df['Capacity Factor - Off Peak'] = pivot_annual_lbo_financials_df['Generation - Off Peak'] / ( pivot_annual_lbo_financials_df['ICAP'] * ( pivot_annual_lbo_financials_df['Hours - Off Peak'] ))

    annual_lbo_financials_df = pd.melt(pivot_annual_lbo_financials_df,
                                       id_vars=['portfolio','scenario','version','entity','year'],
                                       value_vars=[item for item in list(pivot_annual_lbo_financials_df.columns) if item not in ['portfolio','scenario','version','entity','year']],
                                       var_name='fsli',
                                       value_name='value')

    grouped_sum_fslis_lbo_financials_df = annual_lbo_financials_df


    annual_lbo_financials_view_df = pd.DataFrame()
    for entity in entity_list:
        entity_annual_financials_df = grouped_sum_fslis_lbo_financials_df.loc[grouped_sum_fslis_lbo_financials_df.entity == entity]
        pivot_annual_lbo_financials_df = pd.pivot_table(entity_annual_financials_df, index=['portfolio','scenario','version','entity','fsli'], columns=['year'], values='value', aggfunc=np.sum)
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reset_index()
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df[[item for item in pivot_annual_lbo_financials_df.columns if item not in ['portfolio','scenario','version','entity']]]
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.set_index('fsli')

        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reindex(LBO_SUM_FSLIS + ['ICAP','Capacity Factor','Capacity Factor - On Peak','Capacity Factor - Off Peak'])
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reset_index()
        pivot_annual_lbo_financials_df['entity'] = entity
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df[['entity'] + [item for item in pivot_annual_lbo_financials_df.columns if item not in ['entity']]]
        annual_lbo_financials_view_df = annual_lbo_financials_view_df.append(pivot_annual_lbo_financials_df)


    for r in dataframe_to_rows(annual_lbo_financials_view_df, index=False, header=True):
        annual_consolidated_tab.append(r)


    annual_consolidated_tab.title = 'AnnualByPlant'

    # annual view for fsli per portfolio

    annual_fsli_consolidated_tab = wb.copy_worksheet(wb.active)

    sum_fslis_lbo_financials_df = lbo_financials_df.loc[lbo_financials_df.fsli.isin(LBO_SUM_FSLIS)]
    sum_fslis_lbo_financials_df.loc[:,'year'] = pd.DatetimeIndex(sum_fslis_lbo_financials_df['period']).year
    grouped_sum_fslis_lbo_financials_df = sum_fslis_lbo_financials_df.groupby(['portfolio','scenario','version','fsli','year']).sum()
    grouped_sum_fslis_lbo_financials_df = grouped_sum_fslis_lbo_financials_df.reset_index()

    grouped_sum_fslis_lbo_financials_df = grouped_sum_fslis_lbo_financials_df[['fsli','year','value']]

    annual_lbo_financials_view_df = pd.pivot_table(grouped_sum_fslis_lbo_financials_df, index=['fsli'], columns=['year'], values='value', aggfunc=np.sum)

    annual_lbo_financials_view_df = annual_lbo_financials_view_df.reset_index()

    for r in dataframe_to_rows(annual_lbo_financials_view_df, index=False, header=True):
        annual_fsli_consolidated_tab.append(r)

    annual_fsli_consolidated_tab.title = 'AnnualFSLI'

    # monthly view for individual plant
    for entity in entity_list:
        entity_lbo_financials_df = lbo_financials_df.loc[lbo_financials_df.entity == entity]
        entity_tab = wb.copy_worksheet(wb.active)
        entity_tab.title = entity.replace("/"," ")
        entity_df = pd.pivot_table(entity_lbo_financials_df, index=['portfolio','scenario','version','fsli','entity'], columns=['period'], values='value', aggfunc=np.sum)
        entity_df = entity_df.reset_index()
        entity_df = entity_df[[column for column in entity_df.columns if column not in ['portfolio','scenario','version','entity']]]
        entity_df.rename(columns={'fsli': entity}, inplace=True)

        entity_df = entity_df.set_index(entity)

        entity_df = entity_df.reindex(['Energy Revenue','Delivered Fuel Expense','Variable O&M Expense',
                                       'Net Emissions Expense','Gross Energy Margin','Hedges',
                                       'Net Energy Margin','Fixed Fuel Transport','Capacity Revenue',
                                       'Ancillary Revenue','Other Revenue','Gross Margin',
                                       'FOM','Taxes','Insurance','Fixed Costs',
                                       'EBITDA','Capex','EBITDA less Capex',
                                       'ICAP', 'Generation', 'Generation - On Peak', 'Generation - Off Peak',
                                       'Capacity Factor', 'Capacity Factor - On Peak','Capacity Factor - Off Peak',
                                       'Realized Power Price - On Peak', 'Realized Power Price - Off Peak',
                                       'Realized Fuel Price - On Peak', 'Realized Fuel Price - Off Peak',
                                       'Realized Spread - ATC','Realized Spread - Off Peak','Realized Spread - On Peak',
                                       'Hours - On Peak', 'Hours - Off Peak'])




        entity_df = entity_df.reset_index()
        for r in dataframe_to_rows(entity_df, index=False, header=True):
            entity_tab.append(r)



    wb.remove_sheet(wb.active)


    wb.save(dest_file_path + r"\\" + portfolio + "_" + lbo_financials_df['scenario'].iloc[0].replace(portfolio+" ", '') + "_" + lbo_financials_df['version'].iloc[0] + "_lbo_financials.xlsx")




def convert_annual_lbo_financials(lbo_financials_df):

    # step 1, apply lbo color for sinage
    pivot_lbo_financials_df = pd.pivot_table(lbo_financials_df, index=['portfolio','scenario','version','entity','period'], columns=['fsli'], values='value', aggfunc=np.sum)

    for fsli in LBO_FSLI_COLOR:
        pivot_lbo_financials_df[fsli] = pivot_lbo_financials_df[fsli] * LBO_FSLI_COLOR[fsli]

    pivot_lbo_financials_df = pivot_lbo_financials_df.reset_index()

    lbo_financials_df = pd.melt(pivot_lbo_financials_df, id_vars=['portfolio','scenario','version','entity','period'],
                                value_vars=[item for item in list(pivot_lbo_financials_df.columns) if item not in ['portfolio','scenario','version','entity','period']],
                                var_name='fsli',
                                value_name='value')

    lbo_financials_df = lbo_financials_df.reset_index()

    entity_list = list(sorted(list(set(list(lbo_financials_df['entity'])))))



    sum_fslis_lbo_financials_df = lbo_financials_df.loc[lbo_financials_df.fsli.isin(LBO_SUM_FSLIS)]
    sum_fslis_lbo_financials_df.loc[:,'year'] = pd.DatetimeIndex(sum_fslis_lbo_financials_df['period']).year
    grouped_sum_fslis_lbo_financials_df = sum_fslis_lbo_financials_df.groupby(['portfolio','scenario','version','entity','fsli','year']).sum()
    grouped_sum_fslis_lbo_financials_df = grouped_sum_fslis_lbo_financials_df.reset_index()


    average_fslis = ['ICAP']
    average_fslis_lbo_financials_df = lbo_financials_df.loc[lbo_financials_df.fsli.isin(average_fslis)]
    average_fslis_lbo_financials_df.loc[:,'year'] = pd.DatetimeIndex(average_fslis_lbo_financials_df['period']).year
    grouped_average_fslis_lbo_financials_df = average_fslis_lbo_financials_df.groupby(['portfolio','scenario','version','entity','fsli','year']).mean()
    grouped_average_fslis_lbo_financials_df = grouped_average_fslis_lbo_financials_df.reset_index()
    annual_lbo_financials_df = grouped_sum_fslis_lbo_financials_df.append(grouped_average_fslis_lbo_financials_df)


    """ pivot it to per year per column """
    pivot_annual_lbo_financials_df = pd.pivot_table(annual_lbo_financials_df, index=['portfolio','scenario','version','entity','year'], columns=['fsli'], values='value', aggfunc=np.sum)

    pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reset_index()

    pivot_annual_lbo_financials_df['Capacity Factor'] = pivot_annual_lbo_financials_df['Generation'] / ( pivot_annual_lbo_financials_df['ICAP'] * ( pivot_annual_lbo_financials_df['Hours - On Peak'] + pivot_annual_lbo_financials_df['Hours - Off Peak'] ))
    pivot_annual_lbo_financials_df['Capacity Factor - On Peak'] = pivot_annual_lbo_financials_df['Generation - On Peak'] / ( pivot_annual_lbo_financials_df['ICAP'] * ( pivot_annual_lbo_financials_df['Hours - On Peak'] ))
    pivot_annual_lbo_financials_df['Capacity Factor - Off Peak'] = pivot_annual_lbo_financials_df['Generation - Off Peak'] / ( pivot_annual_lbo_financials_df['ICAP'] * ( pivot_annual_lbo_financials_df['Hours - Off Peak'] ))

    annual_lbo_financials_df = pd.melt(pivot_annual_lbo_financials_df,
                                       id_vars=['portfolio','scenario','version','entity','year'],
                                       value_vars=[item for item in list(pivot_annual_lbo_financials_df.columns) if item not in ['portfolio','scenario','version','entity','year']],
                                       var_name='fsli',
                                       value_name='value')

    grouped_sum_fslis_lbo_financials_df = annual_lbo_financials_df


    annual_lbo_financials_view_df = pd.DataFrame()
    for entity in entity_list:
        entity_annual_financials_df = grouped_sum_fslis_lbo_financials_df.loc[grouped_sum_fslis_lbo_financials_df.entity == entity]
        pivot_annual_lbo_financials_df = pd.pivot_table(entity_annual_financials_df, index=['portfolio','scenario','version','entity','fsli'], columns=['year'], values='value', aggfunc=np.sum)
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reset_index()
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df[[item for item in pivot_annual_lbo_financials_df.columns if item not in ['portfolio','scenario','version','entity']]]
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.set_index('fsli')

        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reindex(LBO_SUM_FSLIS + ['ICAP','Capacity Factor','Capacity Factor - On Peak','Capacity Factor - Off Peak'])
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df.reset_index()
        pivot_annual_lbo_financials_df['entity'] = entity
        pivot_annual_lbo_financials_df = pivot_annual_lbo_financials_df[['entity'] + [item for item in pivot_annual_lbo_financials_df.columns if item not in ['entity']]]
        annual_lbo_financials_view_df = annual_lbo_financials_view_df.append(pivot_annual_lbo_financials_df)

    return annual_lbo_financials_df[['portfolio','scenario','version','entity','fsli','year','value']], annual_lbo_financials_view_df



def write_lbo_financials_diff_report(dest_file_path, portfolio, first_lbo_financials_df, second_lbo_financials_df):

    wb = Workbook()

    first_annual_lbo_financials_df, first_annual_lbo_financials_view_df = convert_annual_lbo_financials(first_lbo_financials_df)
    second_annual_lbo_financials_df, second_annual_lbo_financials_view_df = convert_annual_lbo_financials(second_lbo_financials_df)

    merged_annual_lbo_financials_df = pd.merge(first_annual_lbo_financials_df, second_annual_lbo_financials_df, on=['portfolio','entity','fsli','year'], how='inner')

    merged_annual_lbo_financials_df = merged_annual_lbo_financials_df.reset_index()

    merged_annual_lbo_financials_df = merged_annual_lbo_financials_df[['portfolio','entity','fsli','year','value_x','value_y']]

    merged_annual_lbo_financials_df.rename(columns={'value_x':'first_financials_value','value_y':'second_financials_value'}, inplace=True)

    merged_annual_lbo_financials_df['diff_first_minus_second'] = merged_annual_lbo_financials_df['first_financials_value'] - merged_annual_lbo_financials_df['second_financials_value']


    first_scenario_version = first_lbo_financials_df.iloc[0]['scenario'] + "-" + first_lbo_financials_df.iloc[0]['version']
    second_scenario_version = second_lbo_financials_df.iloc[0]['scenario'] + "-" + second_lbo_financials_df.iloc[0]['version']

    first_scenario_tab = wb.copy_worksheet(wb.active)
    for r in dataframe_to_rows(first_annual_lbo_financials_view_df, index=False, header=True):
        first_scenario_tab.append(r)
    first_scenario_tab.title = 'FirstScenarioAnnual'


    second_scenario_tab = wb.copy_worksheet(wb.active)
    for r in dataframe_to_rows(second_annual_lbo_financials_view_df, index=False, header=True):
        second_scenario_tab.append(r)
    second_scenario_tab.title = 'SecondScenarioAnnual'

    """
        entity_annual_financials_df, index=['portfolio','scenario','version','entity','fsli'], columns=['year'], values='value', aggfunc=np.sum
    """
    merged_annual_lbo_financials_df = merged_annual_lbo_financials_df[['portfolio', 'entity', 'fsli', 'year', 'diff_first_minus_second']]

    pivot_diff_lbo_financials_df = pd.pivot_table(merged_annual_lbo_financials_df, index=['portfolio','entity','fsli'], columns=['year'], values='diff_first_minus_second', aggfunc=np.sum)

    pivot_diff_lbo_financials_df = pivot_diff_lbo_financials_df.reset_index()

    pivot_diff_lbo_financials_df = pivot_diff_lbo_financials_df[[item for item in pivot_diff_lbo_financials_df.columns if item != 'portfolio']]

    diff_result_df = pd.DataFrame()

    entity_list = list(sorted(list(set(list(merged_annual_lbo_financials_df['entity'])))))

    for entity in entity_list:
        entity_annual_financials_df = pivot_diff_lbo_financials_df.loc[pivot_diff_lbo_financials_df.entity == entity]
        entity_annual_financials_df = entity_annual_financials_df.set_index('fsli')
        entity_annual_financials_df = entity_annual_financials_df.reindex(LBO_SUM_FSLIS + ['ICAP','Capacity Factor','Capacity Factor - On Peak','Capacity Factor - Off Peak'])
        entity_annual_financials_df = entity_annual_financials_df.reset_index()
        diff_result_df = diff_result_df.append(entity_annual_financials_df)

    diff_result_df = diff_result_df[['entity','fsli'] + [item for item in diff_result_df.columns if item not in ['entity','fsli']]]

    diff_scenario_tab = wb.copy_worksheet(wb.active)
    for r in dataframe_to_rows(diff_result_df, index=False, header=True):
        diff_scenario_tab.append(r)
    diff_scenario_tab.title = 'DiffAnnual'

    wb.remove_sheet(wb.active)
    wb.save(dest_file_path + r"\diff_lbo_financials_" + first_scenario_version + " vs " + second_scenario_version +  ".xlsx")



def write_lbo_graph_report(template_path, lbo_financials_df):
    scenario = lbo_financials_df.iloc[0]['scenario']
    version = lbo_financials_df.iloc[0]['version']
    saved_file_path = r"C:\Users\cliu\Kindle Energy Dropbox\Chang Liu\LBO\requirement_docs\vector_report\LBO Graphs " + scenario + version + ".xlsx"
    wb = opx.load_workbook(template_path)
    input_tab = wb['KEAN LBO Financials']
    annual_financials_df, annual_lbo_financials_view_df = convert_annual_lbo_financials(lbo_financials_df)

    for r in dataframe_to_rows(annual_lbo_financials_view_df, index=False, header=True):
        input_tab.append(r)

    wb.save(saved_file_path)




# #

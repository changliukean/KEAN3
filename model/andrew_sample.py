#TODO: Add check to see if revolver draw necessary
#TODO: Add check to see if need to go below target working capital to make DSC
#TODO: Add Act/360 day_count_factor to utilities
#TODO: Add proper calc of average daily balance for Revolver (lc fees, ununused lines, interest expense)
#TODO: Modify ptd calcs to allow for non-calendar year end payment periods
#TODO: Fix InterestRateSwaps in instruments module - calc correct interest payment
#TODO: Build Interest Expense support report
#TODO: Build PTD support report
#TODO: Add forecasted capex to PTD calculation
#TODO: If have multiple DSRAs, each is only valid for its own instrumnet (pari passu solves it?)
#TODO: Fix ptd cleanup to correct for lack of cash to make full payment


import os
import sys
from pathlib import Path
path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
path_utilities = path + '/utility/'
sys.path.insert(0, path_utilities)
import utilities as utils
import instruments as ins

path_test = path + '/test/'
sys.path.insert(0, path_test)
from lbo_waterfall_scenarios import get_cap_structure, get_waterfall, get_portfolio
from lbo_reports import create_lbo_support_report, create_waterfall_report

from datetime import date, datetime
from dateutil.relativedelta import relativedelta

import pandas as pd
import numpy as np
from scipy.optimize import fsolve

import openpyxl as opx

from collections import namedtuple


class Portfolio:
    def __init__(self, **kwargs):
        self.label = kwargs['label']  #where does this get used?
        self.portfolio_scenario = kwargs['portfolio_scenario']
        self.portfolio_version = kwargs['portfolio_version']
        self.cap_struct_scenario = kwargs['cap_struct_scenario']
        self.cap_struct_version = kwargs['cap_struct_version']
        self.waterfall_scenario = kwargs['waterfall_scenario']
        self.waterfall_version = kwargs['waterfall_version']
        self.close_date = kwargs['close_date'].date()
        self.terminal_date = kwargs['terminal_date'].date()
        try:
            self.yield_curve_date = kwargs['yield_curve_date']
        except:
            pass
        try:
            self.yield_curve_version = kwargs['yield_curve_version']
        except:
            pass
        self.etr = kwargs['effective_tax_rate']
        self.first_payment_date = kwargs['first_payment_date'].date()
        self.periodicity_months_ptd = kwargs['periodicity_months_ptd']

        self.ptd = {}               #dictionary (date, amount)

        #TODO what is better way to encapsulate database connection
        HOST = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com'
        USER = 'Andrew'
        PASSWORD = 'Kindle01'
        DATABASE = 'kean'
        self.cnx = utils.generate_connection_instance(HOST, USER, PASSWORD, DATABASE)
        return

    def get_amount(self, item, period, cash=0, prepayments=0.0):
        if item == 'ptd':
            amount = self.calc_ptd(period, prepayments)
        elif item == 'ptd cleanup':
            amount = self.calc_ptd_cleanup(period, prepayments)
        else:
            print('ERROR in portfolio get_amount, unknown item - ', item)
            sys.exit()
        return amount

    def set_amount(self, item, period, cash_flow):
        if item == 'ptd':
            self.ptd[period] = cash_flow
        elif item == 'ptd cleanup':
            #only necessar for optimization runs, zero when complete
            pass
        else:
            print('ERROR in portfolio set_amount, unknown item - '. item)
            sys.exit()
        return

    def is_payment_date(self, period):
        if period < self.first_payment_date:
            result = False
        elif period > utils.calc_next_month_end(self.terminal_date, 'date', self.periodicity_months_ptd):
            result = False
        elif (period.year * 12 + period.month - self.first_payment_date.year * 12
              - self.first_payment_date.month) % self.periodicity_months_ptd == 0:
            result = True
        else:
            result = False
        return result

    def calc_number_ptd_payments_remaining(self, period):
        if period.year == self.terminal_date.year:
            number_ptd_payments = int((self.terminal_date.month - period.month) / self.periodicity_months_ptd) + 1
        else:
            number_ptd_payments = int((12 - period.month) / self.periodicity_months_ptd) + 1
        return number_ptd_payments

    def calc_ptd(self, period, prepayments, flag_cleanup=False):
        # ) check if ptd payment date (usually quarterly)
        #1) get OpCo CFO
        #3) calc interest expense
        #   - actual & forecast
        #   - oid & dfc
        #4) get tax depreciation
        #   - get capex
        #   - get tax register (or variation)
        #   - calc tax depreciation
        #5) get effective tax rate
        #6) determine how much gets paid this period
        #   - determine how many periods (acquisition different than existing)
        #   - determine how many periods already paid (subtract from annual calc)
        # REUSE FUNCTION FOR CLEANUP CALC
        #   - NEED CLEANUP FLAG TO NOT INCLUDE CURRENT PERIOD PTD

        number_payments = self.calc_number_ptd_payments_remaining(period)
        ptd = 0.0
        if self.is_payment_date(period):
            month_number = (period.year * 12 + period.month) - (self.close_date.year * 12 + self.close_date.month)
            if period.year == self.close_date.year:
                months_in_tax_year = 12 - self.close_date.month + 1      #assumes close date is start of month
            elif period.year == self.terminal_date.year:
                months_in_tax_year = self.terminal_date.month           #assumes terminal date is last day of month
            else:
                months_in_tax_year = 12
            tax_ebitda = 0.0
            tax_capex = 0.0
            for ins in portfolio.instruments:
                if self.instruments[ins].type == 'OperatingCompany':
                    ebitda = self.instruments[ins].ebitda           #dictionary of cash flows
                    capex = self.instruments[ins].capex           #dictionary of cash flows
                    for cashflow in ebitda:
                        if cashflow.year == period.year and cashflow >= self.close_date:
                            tax_ebitda += ebitda[cashflow]
                    for cashflow in capex:
                        if cashflow.year == period.year and cashflow >= self.close_date:
                            tax_capex += capex[cashflow]
                    tax_depreciation = -self.instruments[ins].tax_register.calc_depreciation(period)
                if self.instruments[ins].type in ['Debt', 'FixedDebt', 'FloatingDebt', 'MezzanineDebt']:
                    cash_interest = self.instruments[ins].calc_tax_interest(period, prepayments, month_number, months_in_tax_year)
            ptd_paid = 0.0
            for cashflow in self.ptd:
                if cashflow.year == period.year:
                    if flag_cleanup and cashflow == period:
                        pass
                    else:
                        ptd_paid += self.ptd[cashflow]
            ptd_tax_year = -(tax_ebitda + cash_interest + tax_depreciation) * self.etr
            ptd = (ptd_tax_year - ptd_paid) / number_payments
        return ptd

    def calc_ptd_cleanup(self, period, prepayments):
        #change the solver prepayments to reflect the current period actual prepayment,
        #   then call calc_ptd
        flag_cleanup = True
        ptd_cleanup = 0.0
        if self.is_payment_date:
            month_number = (period.year * 12 + period.month) - (self.close_date.year * 12 + self.close_date.month)
            for ins in portfolio.instruments:
                if self.instruments[ins].type in ['Debt', 'FixedDebt', 'FloatingDebt']:
                    if self.instruments[ins].flag_prepayable:
                        prepayments[month_number] = self.instruments[ins].prepayments[period]
            ptd = self.calc_ptd(period, prepayments, flag_cleanup)
            ptd_cleanup = ptd - self.ptd[period]
        #print('ptd cleanup calc = ', ptd, self.ptd[period], ptd_cleanup)
        return ptd_cleanup


class OperatingCompany:
    # TODO figure out how to laod CFO from database
    def __init__(self, **kwargs):
        self.type = kwargs['class']
        self.label = kwargs['label']
        self.working_capital = kwargs['working_capital'] / UNITS
        self.working_capital_target = kwargs['working_capital_target'] / UNITS
        self.interest_rate_wc = kwargs['interest_rate_wc']
        self.periodicity_months = kwargs['periodicity_months']
        self.day_count = kwargs['day_count']
        self.scenario_date_start = kwargs['scenario_date_start'].date()
        self.scenario_date_end = kwargs['scenario_date_end'].date()
        try:
            self.financials_scenario = kwargs['financials_scenario']
            self.financials_version = kwargs['financials_version']
            self.financials_company = kwargs['financials_company']
            self.financials_entity = kwargs['financials_entity']
        except:
            pass
        self.flag_tax_asset_detail = kwargs['flag_tax_asset_detail']

        self.cfo = self.get_cfo()                   # dictionay (date/amount)
        self.ebitda = self.get_ebitda()             # dictionay (date/amount)
        self.capex = self.get_capex()               # dictionay (date/amount)
        self.get_tax_register()

        # for reporting
        self.working_capital_history = {}           # dictionary (date/amount)
        self.cfo_history = {}
        self.interest_income_history = {}
        self.working_capital_change = {}

    def get_cfo(self):
        # TODO: allow for specific companry and entity in query
        # use dict initially, consider dataframe on refactor
        if hasattr(self, 'financials_scenario'):
            # get CFO from database
            query = ("SELECT period, sum(value) as value FROM financials WHERE scenario = %s AND version = %s AND "
                     "account = 'EBITDA less Capex' GROUP BY period")
            df_cfo = pd.read_sql(query, cnx, params=(self.financials_scenario, self.financials_version), index_col=['period'])
            df_cfo['value'] = df_cfo['value'] / UNITS
            cfo = df_cfo.to_dict()['value']
        return cfo

    def get_ebitda(self):
        #TODO: allow for specific companry and entity in query
        #use dict initially, consider dataframe on refactor
        if hasattr(self, 'financials_scenario'):
            #get CFO from database
            query = ("SELECT period, sum(value) as value FROM financials WHERE scenario = %s AND version = %s AND "
                     "account = 'EBITDA' GROUP BY period")
            df_ebitda = pd.read_sql(query, cnx, params=(self.financials_scenario, self.financials_version), index_col=['period'])
            df_ebitda['value'] = df_ebitda['value'] / UNITS
            ebitda = df_ebitda.to_dict()['value']
        else:
            print('ERROR - no financials selected for OpCo get_ebitda')
            sys.exit()
        return ebitda

    def get_capex(self):
        #TODO: allow for specific companry and entity in query
        #use dict initially, consider dataframe on refactor
        if hasattr(self, 'financials_scenario'):
            #get CFO from database
            query = ("SELECT period, sum(value) as value FROM financials WHERE scenario = %s AND version = %s AND "
                     "account in ('Maintenance Capex', 'Environmental Capex', 'LTSA Capex', 'Growth Capex') GROUP BY period")
            df_capex = pd.read_sql(query, cnx, params=(self.financials_scenario, self.financials_version), index_col=['period'])
            df_capex['value'] = df_capex['value'] / UNITS
            capex = df_capex.to_dict()['value']
        else:
            print('ERROR - no financials selected for OpCo get_ebitda')
            sys.exit()
        return capex

    def get_tax_register(self):
        if self.flag_tax_asset_detail:
            self.tax_register = TaxRegister(self.label)
            tax_assets = get_tax_register_from_xlsx()
            for asset in tax_assets:
                self.tax_register.add_asset(FixedAsset(*asset))
        return

    def get_amount(self, metric, period, cash=0, prepay=0.0):
        if metric == 'CFO':
            try:
                #period_end = min(period, self.scenario_date_end)
                #day_count_factor = utils.calc_day_count_factor(self.day_count, self.calc_prior_payment_period(period), period_end)
                amount = self.cfo[period]
            except Exception as e:
                print("ERROR - invalid date for CFO ", period)
        elif metric == 'working capital':
            amount = max(self.working_capital, 0)
        elif metric == 'interest income':
            amount = self.calc_interest_income(period)
        elif metric == 'sweep':
            #should be last item in waterfall; happens on non-quarter end months; cash sits in bank account
            #self.working_capital += cash
            amount = -cash
        elif metric == 'working capital reset':
            amount = max(-self.working_capital_target, -cash)
        else:
            print("Error in OperatingCompany get_amount - unknown metric ", metric)
            sys.exit()
        return amount

    def set_amount(self, item, period, cash_flow):
        if item == 'CFO':
            self.cfo_history[period] = cash_flow
        elif item == 'interest income':
            self.interest_income_history[period] =  cash_flow
        elif item == 'working capital':
            self.working_capital_history[period] = self.working_capital
            self.working_capital_change[period] = cash_flow
            self.working_capital -= cash_flow
            #self.working_capital_history.append((period, self.working_capital))
        elif item == 'working capital reset':
            self.working_capital -= cash_flow
        elif item == 'sweep':
            self.working_capital -= cash_flow
        else:
            print('Error - unknown item in OperatingCompany set_amount ', item)
        return

    def calc_prior_payment_period(self, period):
        if period <= self.scenario_date_start:
            prior_period = None
        elif period > self.scenario_date_end:
            #check if stub final period
            if utils.calc_next_month_end(self.scenario_date_end, 'date', self.periodicity_months) < period :
                prior_period = None
            else:
                prior_period = utils.calc_next_month_end(period, 'date', -self.periodicity_months)
        elif period < utils.calc_next_month_end(self.scenario_date_start, 'date', +self.periodicity_months):
            prior_period = self.scenario_date_start
        else:
            prior_period = utils.calc_next_month_end(period, 'date', -self.periodicity_months)
        return prior_period

    def calc_interest_income(self, period):
        #calculates interest income on working capital
        period_end = min(period, self.scenario_date_end)
        day_count_factor = utils.calc_day_count_factor(self.day_count, self.calc_prior_payment_period(period), period_end)
        interest = self.working_capital * self.interest_rate_wc * day_count_factor
        return interest


class Revolver:
    def __init__(self, **kwargs):
        self.type = kwargs['class']
        self.label = kwargs['label']
        self.issue_date = kwargs['issue_date']
        self.term = kwargs['term']
        self.maturity_date = self.issue_date + relativedelta(months=+self.term) + relativedelta(days=-1)
        self.maturity_date = self.maturity_date.date()
        self.credit_line = kwargs['credit_line'] / UNITS
        self.initial_balance = kwargs['initial_balance'] / UNITS
        self.index_name = kwargs['index_name']
        self.margin = kwargs['margin']
        self.day_count = kwargs['day_count']
        self.periodicity_months = kwargs['periodicity_months']
        self.undrawn_line_fee = kwargs['undrawn_line_fee']
        try:
            self.dsra = kwargs['dsra'] / UNITS
        except:
            pass
        self.dsra_months = kwargs['dsra_months']
        self.first_payment_date = kwargs['first_payment_date'].date()
        try:
            self.letters_of_credit = kwargs['letters_of_credit'] / UNITS
        except:
            self.letters_of_credit = 0.0
        try:
            self.lc_fee_rate = kwargs['lc_fee_rate']
        except:
            self.lc_fee_rate = 0.0
        self.principal = self.initial_balance
        self.set_index()
        self.line_fees = {}
        self.lc_fees = {}
        self.interest_expense ={}
        self.dsra_change = {}
        self.dsra_release = {}
        self.draws = {}
        self.sweeps = {}

    def set_index(self):
        #pull libor curve from KEAN
        self.index = self.get_adj_libor()
        return

    def get_libor(self):
        #purpose: return df of monthly libor rates, these have various forward dates
        #assume LIBOR-1MO initially
        #TO DO: allow different scenarios and versions
        query = ("SELECT period, price FROM prices WHERE scenario = 'Actuals' AND version = %s "
                 "AND instrument_id = %s AND valuation_date = %s ORDER BY period")
        df = pd.read_sql(query, cnx, params=(YIELD_CURVE_VERSION, self.index_name, YIELD_CURVE_DATE))
        return df

    def get_adj_libor(self):
        #purpose: convert df from get_libor to curve based on month end dates
        #call get_libor, interpolate/extropolate to month_end data points
        #TODO: overload start and end date to allow extrapolation of rates
        df = self.get_libor()
        period = utils.calc_month_end(df['period'].min(), 'date')
        curve = {}
        while period < df.iloc[0]['period']:
            #extropolate backwards - should never happen
            increment = (df.iloc[1]['price'] - df.iloc[0]['price']) / (df.iloc[1]['period'] - df.iloc[0]['period']).days
            interval = (df.iloc[0]['period'] - period).days
            curve[period] = df.iloc[0]['price'] - interval * increment
            period = utils.calc_next_month_end(period, 'date')
        while period <= df['period'].max():
            #interpolate
            bottom_date = max(df.loc[(df['period']<=period)]['period'])
            bottom_yield = df.loc[df['period']==bottom_date]['price'].values[0]
            if period == bottom_date:
                curve[period] = bottom_yield
            elif df.loc[df['period']>period].shape[0] == 0:
                #need to extropolate - does not happen unless overload start and end dates
                increment = (df.iloc[-1]['price'] - df.iloc[-2]['price']) / ((df.iloc[-1]['period'] - df.iloc[-2]['period']).days)
                interval = (period - df.iloc[-1]['period']).days
                curve[period] = df.iloc[-1]['price'] + interval * increment
            else:
                top_date = min(df.loc[(df['period']>=period)]['period'])
                bottom_yield = df.loc[df['period']==bottom_date]['price'].values[0]
                top_yield = df.loc[df['period']==top_date]['price'].values[0]
                increment = (top_yield - bottom_yield) / (top_date - bottom_date).days
                interval = (period - bottom_date).days
                curve[period] = bottom_yield + interval * increment
            period = utils.calc_next_month_end(period, 'date')
            #df_curve = pd.DataFrame(curve, columns= ['period', 'libor'])
        return curve

    def get_amount(self, item, period, cash=0, prepay=0.0):
        if item == 'undrawn line fee':
            if self.is_payment_date(period):
                try:
                    amount = self.calc_undrawn_line_fee(period)
                except:
                    print("ERROR - invalid date for Revolver ", period)
                    sys.exit()
            else:
                amount = 0.0
        elif item == 'lc fees':
            if self.is_payment_date(period):
                amount = self.calc_lc_fees(period)
            else:
                amount = 0.0
        elif item == 'draw':
            amount = self.credit_line - self.principal - self.letters_of_credit
        elif item == 'interest expense':
            if self.is_payment_date(period):
                amount = -self.calc_interest_expense(period)
            else:
                amount = 0.0
        elif item == 'dsra reset':
            #amount = self.calc_dsra_change(period)
            amount = 0.0
        elif item == 'dsra release':
            #placeholder
            amount = 0.0
        elif item == 'sweep':
            amount = -(self.principal)
        else:
            print("Error in Revolver get_amount - unknown metric ", metric)
            sys.exit()
        return amount

    def set_amount(self, item, period, cash_flow):
        if item == 'undrawn line fee':
            self.line_fees[period] = cash_flow
        elif item == 'lc fees':
            self.lc_fees[period] = cash_flow
        elif item == 'interest expense':
            self.interest_expense[period] = cash_flow
        elif item == 'dsra change':
            self.dsra_change[period] = cash_flow
        elif item == 'dsra release':
            self.dsra_release[period] = cash_flow
        elif item == 'draw':
            self.draws[period] = cash_flow
            self.principal += cash_flow
        elif item == 'sweep':
            self.sweeps[period] = cash_flow
            self.principal += cash_flow
        else:
            print('Error - unknown item in Revolver set_amount ', item)
        return

    def is_payment_date(self,period):
        if period < self.first_payment_date:
            result = False
        elif period > utils.calc_next_month_end(self.maturity_date, 'date', self.periodicity_months):
            result = False
        elif (period.year * 12 + period.month - self.first_payment_date.year * 12
              - self.first_payment_date.month) % self.periodicity_months == 0:
            result = True
        else:
            result = False
        return result

    def calc_prior_payment_period(self, period):
        #only gets called if valid payment date
        #need to check if first payment date
        if period == self.first_payment_date:
            prior_period = self.issue_date
        else:
            prior_period = utils.calc_next_month_end(period, 'date', -self.periodicity_months)
        return prior_period

    def calc_interest_rate(self, period):
        try:
            self.interest_rate = self.index[period] + self.margin
        except:
            print("Error in calc_interest_rate - invalid period ", period)
        return

    def calc_interest_expense(self, period):
        #calculates interest income on working capital
        day_count_factor = utils.calc_day_count_factor(self.day_count, self.calc_prior_payment_period(period), period)
        self.calc_interest_rate(period)
        interest = self.calc_principal_bop(period) * self.interest_rate * day_count_factor
        return interest

    def calc_dsra(self, period):
        #initially assume no paydown of debt (removes circularity of calc)
        #TODO include paydown of debt
        #1) LC fees
        lc_fees = 0.0 * self.dsra_months / 12
        #2) undrawn line fee
        undrawn_line_fee = (self.credit_line - self.principal) * self.undrawn_line_fee * self.dsra_months / 12
        #3) interest expense
        interest_expense = 0.0
        if self.dsra_months < self.periodicity_months:
            #this is necessary for annual models with 6 month dsra requirements
            #need to determine what correct period to call calc_interest
            #next_period = utils.calc_next_month_end(period, 'date', self.periodicity_months)
            #interest = self.calc_interest_expense(next_period, self.principal - prepayment)
            #interest_portion = self.dsra_months / self.periodicity_months * interest
            pass
        else:
            if period < self.first_payment_date:
                #determine initial stub period
                #   after first payment, should only check dsra on payment date
                #   initially assume stub payment index = first payment index
                #TODO: calc proper stub index rate
                day_count_factor = utils.calc_day_count_factor(self.day_count, PORTFOLIO_START_DATE, self.first_payment_date)
                interest_expense += self.initial_balance * (self.index[period] + self.margin) * day_count_factor
                #determine how many whole payment periods follow
                #   assumes month of close counts as 1 month
                dsra_end = utils.calc_next_month_end(PORTFOLIO_START_DATE, 'date', self.dsra_months - 1)
                pmt_periods = int((dsra_end.year * 12 + dsra_end.month - PORTFOLIO_START_DATE.year * 12 -
                                   PORTFOLIO_START_DATE.month)/self.periodicity_months)
                current_period = self.first_payment_date
                next_period = utils.calc_next_month_end(self.first_payment_date, 'date', self.periodicity_months)
                for i in range(pmt_periods):
                    day_count_factor = utils.calc_day_count_factor(self.day_count, current_period, next_period)
                    #assume no paydown in balance
                    interest_expense += self.initial_balance * (self.index[current_period] + self.margin) * day_count_factor
                    current_period = next_period
                    next_period = utils.calc_next_month_end(next_period, 'date', self.periodicity_months)
                #check if stub end period
                stub_months = ((dsra_end.year * 12 + dsra_end.month - PORTFOLIO_START_DATE.year * 12 -
                                   PORTFOLIO_START_DATE.month) % self.periodicity_months)
                if stub_months != 0:
                    day_count_factor = utils.calc_day_count_factor(self.day_count, current_period, next_period)
                    #assume no paydown in balance
                    interest_expense += self.initial_balance * (self.index[current_period] + self.margin) * day_count_factor
            else:
                #normal dsra calc on a payment period
                while period < utils.calc_next_month_end(period, 'date', self.dsra_months):
                    next_period = utils.calc_next_month_end(period, 'date', self.periodicity_months)
                    day_count_factor = utils.calc_day_count_factor(self.day_count, period, next_period)
                    interest_expense += self.principal * (self.index[period] + self.margin) * day_count_factor
                    period = next_period
        return lc_fees + undrawn_line_fee + interest_expense

    def calc_dsra_change(self,period):
        #figure out day count factor implications at later time
        interest = 0.0
        undrawn_line_fee = (self.credit_line - self.principal) * self.undrawn_line_fee
        lc_fees = 0.0
        dsra_new = (interest + lc_fees + undrawn_line_fee) * self.dsra_months / 12
        return self.dsra - dsra_new

    def calc_undrawn_line_fee(self, period):
        if self.is_payment_date(period):
            period_end = min(period, self.maturity_date)
            day_count_factor = utils.calc_day_count_factor(self.day_count, self.calc_prior_payment_period(period), period_end)
            amount = -(self.credit_line - self.principal - self.letters_of_credit + self.draws[period]) * self.undrawn_line_fee * day_count_factor
        else:
            amount = 0.0
        return amount

    def calc_lc_fees(self, period):
        period_end = min(period, self.maturity_date)
        day_count_factor = utils.calc_day_count_factor(self.day_count, self.calc_prior_payment_period(period), period_end)
        amount = -self.letters_of_credit * self.lc_fee_rate * day_count_factor
        return amount


    def calc_principal_bop(self, period):
        #necessary for interest expense calc
        draws = 0.0
        sweeps = 0.0
        period_loop = utils.calc_month_end(self.issue_date, 'date')
        while period_loop < period:
            try:
                draws += self.amortization[period_loop]
            except:
                draws += 0.0
            try:
                sweeps += self.prepayments[period_loop]
            except:
                sweeps += 0.0
            period_loop = utils.calc_next_month_end(period_loop, 'date')
        principal = self.initial_balance + draws + sweeps
        return principal


class Debt:
    def __init__(self, **kwargs):
        #self.name = name
        self.type = kwargs['class']
        self.label = kwargs['label']
        self.issue_date = kwargs['issue_date'].date()
        self.initial_balance = kwargs['initial_balance'] / UNITS
        self.annual_amort_percent = kwargs['annual_amort_percent']
        self.interest_date_start = kwargs['interest_date_start'].date()
        self.amort_date_start = kwargs['amort_date_start'].date()
        self.periodicity_months = kwargs['periodicity_months']
        #self.set_periodicity_months()
        self.amort_const = (self.initial_balance * self.annual_amort_percent / (12 / self.periodicity_months))
        self.day_count = kwargs['day_count']
        self.sweep_percent = kwargs['sweep_percent']
        self.term = kwargs['term']
        self.maturity_date = self.issue_date + relativedelta(months=+self.term) + relativedelta(days=-1)
        self.oid = kwargs['oid']
        self.dfc = kwargs['dfc']
        self.flag_prepay_offset = kwargs['flag_prepay_offset']
        self.dsra_months = kwargs['dsra_months']
        self.dsra = self.initialize_dsra()
        self.dsra_interest_rate = kwargs['dsra_interest_rate']
        self.flag_prepayable = kwargs['flag_prepayable']
        try:
            self.flag_swaps = kwargs['flag_swaps']
            self.company = kwargs['company']
        except:
            self.flag_swaps = False
        #self.lc_fees = kwargs['lc_fees']
        self.principal = self.initial_balance
        self.amortization = {}
        self.prepayments = {}
        self.interest_expense = {}
        self.principal_history_bop = {}
        self.principal_history_eop = {}
        self.dsra_change = 0.0
        self.prepayment = 0.0
        self.cfas_flag = False

    def set_periodicity_months(self):
        if self.periodicity == 'monthly':
            self.periodicity_months = 1
        elif self.periodicity == 'quarterly':
            self.periodicity_months =3
        elif self.periodicity == 'semiannual':
            self.periodicity_months = 6
        elif self.periodicity == 'annual':
            self.periodicity_months = 12
        else:
            print('ERROR: unknown periodicity in DebtInstrument ini - ', self.periodicity)
        return

    def set_amount(self, item, period, cash_flow):
        if item == 'interest income':
            #not clear if anything needs to happen
            pass
        elif item == 'dsra release':
            self.dsra -= cash_flow
        elif item == 'interest expense':
            self.interest_expense[period] = cash_flow
        elif item == 'amortization':
            self.principal_history_bop[period] = self.principal
            self.amortization[period] = cash_flow
            self.principal += cash_flow
        elif item == 'dsra reset':
            self.dsra -= cash_flow
        elif item == 'sweep':
            if self.flag_prepayable:
                self.prepayments[period] =  cash_flow
                self.principal += cash_flow
                self.principal_history_eop[period] = self.principal
        elif item == 'dsra cleanup':
            #self.dsra -= cash_flow
            pass
        else:
            print("Error - unknow item sent to set_amount ", item)
        return

    def initialize_dsra(self):
        #test if initial dsra balance is loaded with debt profile
        try:
            self.dsra = kwargs['dsra_months']
        except:
            pass
        #if initial dsra balance is not loaded with debt profile, calculate
        if not hasattr(self, 'dsra'):
            months_to_first_payment = (self.interest_date_start.year * 12 + self.interest_date_start.month -
                utils.calc_next_month_end(self.issue_date, 'date', -1).year * 12 -
                utils.calc_next_month_end(self.issue_date, 'date', -1).month)
            #initialize values
            dsra_princ = self.dsra_months / 12 * self.annual_amort_percent * self.initial_balance
            dsra_int = 0.0
            principal = self.initial_balance
            prior_period = self.issue_date
            if months_to_first_payment % self.periodicity_months == 0:
                #no stub period
                period = utils.calc_next_month_end(utils.calc_next_month_end(self.issue_date, 'date', -1), 'date', self.periodicity_months)
                #print(principal, prior_period, period, self.dsra_months / self.periodicity_months)
                #sys.exit()
                for i in range(int(self.dsra_months / self.periodicity_months)):
                    interest_rate = self.calc_interest_rate(period)
                    day_count_factor = utils.calc_day_count_factor(self.day_count, prior_period, period)
                    #print(i, interest_rate, day_count_factor, principal)
                    dsra_int += principal * day_count_factor * interest_rate
                    if period >= self.amort_date_start:
                        principal -= self.amort_const
                    prior_period = period
                    period = utils.calc_next_month_end(period, 'date', self.periodicity_months)
            else:
                #has stub periods
                #calc interest for initial stub (add one since issue dates are assumed to be first of month)
                stub_months = (self.interest_date_start.year * 12 + self.interest_date_start.month -
                    self.issue_date.year * 12 - self.issue_date.month + 1) % self.periodicity_months
                period = utils.calc_next_month_end(utils.calc_next_month_end(self.issue_date, 'date', -1), 'date', stub_months)
                interest_rate = self.calc_interest_rate(period)
                day_count_factor = utils.calc_day_count_factor(self.day_count, prior_period, period)
                dsra_int += principal * day_count_factor * interest_rate
                #calc interest expense for middle, normal periods
                if period >= self.amort_date_start:
                    principal -= self.amort_const
                prior_period = period
                period = utils.calc_next_month_end(period, 'date', self.periodicity_months)
                for i in range(int((self.dsra_months-stub_months)/self.periodicity_months)):
                    interest_rate = self.calc_interest_rate(period)
                    day_count_factor = utils.calc_day_count_factor(self.day_count, prior_period, period)
                    dsra_int += principal * day_count_factor * interest_rate
                #calc interest expense for final stub period
                if period >= self.amort_date_start:
                    principal -= self.amort_const
                prior_period = period
                period = utils.calc_next_month_end(utils.calc_next_month_end(self.issue_date, 'date', -1), 'date', self.dsra_months)
                #need to get interest rate assuming normal period end
                interest_rate = self.calc_interest_rate(utils.calc_next_month_end(prior_period, 'date', self.periodicity_months))
                day_count_factor = utils.calc_day_count_factor(self.day_count, prior_period, period)
                dsra_int += principal * day_count_factor * interest_rate
            dsra = self.amort_const * 2 + dsra_int
        return dsra

    def is_interest_payment_date(self,period):
        if period < self.interest_date_start:
            result = False
        elif period > utils.calc_next_month_end(self.maturity_date, 'date', self.periodicity_months):
            result = False
        elif (period.year * 12 + period.month - self.interest_date_start.year * 12
              - self.interest_date_start.month) % self.periodicity_months == 0:
            result = True
        else:
            result = False
        return result

    def calc_amort(self, period):
        if self.is_interest_payment_date(period):
            amount = self.annual_amort_percent * self.initial_balance * self.periodicity_months / 12
        else:
            amount = 0.0
        return amount

    def calc_date_prior_interest_payment(self, period):
        #necessary to calculate number of days in current interest period
        #only gets called if (date_diff.months + date_diff.years * 12) % 3 == 0
        #   so one less * periodicity_months should equal months to prior payment
        # need to test for first payment
        date_diff = relativedelta(period, self.payment_date_start)
        payment_number = (date_diff.months + date_diff.years * 12) % self.periodicity_months
        prior_payment_period = utils.calc_month_end(period + relativedelta(months=-self.periodicity_months), 'date')
        if prior_payment_period < self.issue_date:
            prior_payment_period = self.issue_date
        return prior_payment_period

    def calc_interest_rate(self, period):
        return self.interest_rate

    def calc_interest_expense(self, period, principal=None):
        if principal == None:
            principal = self.principal
        if self.is_interest_payment_date(period):
            prior_period = self.calc_prior_payment_period(period)
            day_count_factor = utils.calc_day_count_factor(self.day_count, prior_period, period)
            int_exp = principal * self.calc_interest_rate(period) * day_count_factor
        else:
            int_exp = 0.0
        if self.flag_swaps:
            swap_payment = ins.calc_swaps_payment(self.company, period, YIELD_CURVE_DATE) / UNITS
        else:
            swap_payment = 0.0
        return int_exp + swap_payment

    def calc_period_days(self, period):
        #return number of days in period for interest calc
        prior_payment_period = self.calc_date_prior_interest_payment(period)
        return (period - prior_payment_period).days

    def calc_next_period(self, period):
        #assumes period passed is a current payment period
        next_period = utils.calc_next_month_end(period, 'date', self.periodicity_months)
        return next_period

    def calc_prior_payment_period(self, period):
        #this function assumes it is called from valid payment date
        if period <= self.issue_date:
            prior_period = None
        elif period > utils.calc_next_month_end(self.maturity_date, 'date', self.periodicity_months):
            prior_period = None
        elif period <= self.interest_date_start:
            prior_period = self.issue_date
        else:
            prior_period = utils.calc_next_month_end(period, 'date', -self.periodicity_months)
        return prior_period

    def calc_dsra_int_inc(self, period):
        day_count_factor = utils.calc_day_count_factor(self.day_count, self.calc_prior_payment_period(period), period)
        return self.dsra * self.dsra_interest_rate * day_count_factor

    def calc_cfas(self, cfas, *args):
        #this functions solves for cfas and dsra_change simultaneously
        #   sets the dsra_change attribute to record results
        #TODO add PTD
        period, cash = args
        prepay = self.sweep_percent * cfas[0]
        principal_eop = self.principal - prepay
        self.dsra_change = self.calc_dsra(period, principal_eop) - self.dsra
        excess_cash = cash + self.dsra_change - cfas
        return excess_cash

    def calc_interest_income(self, period):
        #calc interest income on dsra balances
        if self.dsra == None or self.dsra == 0:
            interest = 0.0
        elif period >= utils.calc_next_month_end(self.maturity_date, 'date', self.periodicity_months):
            interest = 0.0
        else:
            period_end = min(period, self.maturity_date)
            prior_period = self.calc_prior_payment_period(period)
            day_count_factor = utils.calc_day_count_factor(self.day_count, prior_period, period_end)
            interest = self.dsra * self.dsra_interest_rate * day_count_factor
        return interest

    def calc_dsra_change(self, period, cash):
        #assumes 6 month dsra requirement
        #TODO refactor to allow different dsra terms
        cfas = cash
        cfas = fsolve(self.calc_cfas, cfas, (period, cash))[0]
        return self.dsra_change

    def calc_dsra(self, period, principal=None):
        #Returns a positive amount for the required balance of the DSRA
        #NOTE: interest rate is the rate applicable for payment made at period date
        #   if quarterly LIBOR, rate will be 90-day LIBOR ENDING on period date

        #calc principal portion
        if principal == None:
            principal = self.principal
        principal_portion = self.dsra_months / self.periodicity_months * self.amort_const
        #calc interest portion
        months_from_prior_payment = (period.year * 12 + period.month - self.interest_date_start.year * 12
              - self.interest_date_start.month) % self.periodicity_months
        next_period = utils.calc_next_month_end(period, 'date', (self.periodicity_months - months_from_prior_payment))
        next_period_2 = utils.calc_next_month_end(next_period, 'date', self.periodicity_months)
        interest_portion_1 = (principal * self.calc_interest_rate(next_period) *
                              utils.calc_day_count_factor(self.day_count, period, next_period))
        interest_portion_2 = ((principal - self.amort_const) * self.calc_interest_rate(next_period_2) *
                               utils.calc_day_count_factor(self.day_count, next_period, next_period_2))
        dsra = principal_portion + interest_portion_1 + interest_portion_2
        #if self.label == 'Test TLC' and period in [date(2019,12,31), date(2020,3,31), date(2020,6,30)]:
        #    print('DSRA calc = ', next_period, next_period_2, interest_portion_1, interest_portion_2, principal_portion)
        #    print('DSRA interest calc = ', principal, self.calc_interest_rate(next_period), self.calc_interest_rate(next_period_2), utils.calc_day_count_factor(self.day_count, period, next_period))
        #    print(dsra)
        #    #sys.exit()
        return dsra

    def calc_principal_bop(self, period):
        amort = 0.0
        prepay = 0.0
        period_loop = utils.calc_month_end(self.issue_date, 'date')
        while period_loop < period:
            #need try loop as TLC does not have amortization thus no step in the waterfall
            try:
                amort += self.amortization[period_loop]
            except:
                amort += 0.0
            try:
                prepay += self.prepayments[period_loop]
            except:
                prepay += 0.0
            period_loop = utils.calc_next_month_end(period_loop, 'date')
        principal = self.initial_balance + amort + prepay
        return principal

    def calc_tax_interest(self, period, prepayments, month_number, months_in_tax_year):
        actual = sum(self.interest_expense.values())
        forecast = 0.0
        tax_month = utils.calc_next_month_end(period, 'date')
        principal_bop = self.calc_principal_bop(period) - self.calc_amort(period)
        while tax_month <= date(period.year, 12, 31):
            forecast -= self.calc_interest_expense(tax_month, principal_bop + prepayments[month_number])
            principal_bop -= self.calc_amort(tax_month)
            tax_month = utils.calc_next_month_end(tax_month, 'date')
        oid = -((100 - self.oid)/100 * self.initial_balance) / self.term * months_in_tax_year
        dfc = -self.dfc * self.initial_balance / self.term * months_in_tax_year
        interest_expense = actual + forecast + oid + dfc
        return interest_expense


class FixedDebt(Debt):
    #TODO move self.interest rate to Debt class
    #   need to determine if need FixedDebt class
    def __init__(self, **kwargs):
        self.interest_rate = kwargs['interest_rate']
        Debt.__init__(self, **kwargs)
        #TODO add local function to correctly calc initial dsra requirement
        #self.dsra = self.calc_dsra(self.issuance_date)

    def get_amount(self, item, period, cash=0, prepay=0.0):
        if item == 'amortization':
            amount =  -self.calc_amort(period)
        elif item == 'interest expense':
            amount =  -self.calc_interest_expense(period)
        elif item == 'sweep':
            cfas = 0
            sweep = fsolve(self.calc_cfas, cfas, (period, cash))[0]
            amount = -sweep
        elif item == 'dsra change':
            dsra_change = self.calc_dsra_change(period)
            amount = dsra_change
        elif item == 'interest income':
            amount = self.calc_interest_income(period)
        elif item == 'dsra release':
            #placeholder - replace with test to see if dsra needed to make interest
            #   and amoritzation payments
            amount =  0.0
        else:
            print("Error in FixedDebt get_amount - unknown metric ", metric)
            sys.exit()
        return amount


class FloatingDebt(Debt):
    def __init__(self, **kwargs):
        self.margin = kwargs['margin']
        self.index_name = kwargs['index_name']
        self.set_index()
        Debt.__init__(self, **kwargs)

    def set_index(self):
        #pull libor curve from KEAN
        self.index = self.get_adj_libor()
        return

    def get_libor(self):
        #purpose: return df of monthly libor rates, these have various forward dates
        #assume LIBOR-1MO initially
        #TO DO: allow different scenarios and versions
        query = ("SELECT period, price FROM prices WHERE scenario = 'Actuals' AND version = %s "
                 "AND instrument_id = %s AND valuation_date = %s ORDER BY period")
        df = pd.read_sql(query, cnx, params=(YIELD_CURVE_VERSION, self.index_name, YIELD_CURVE_DATE))
        return df

    def get_adj_libor(self):
        #purpose: convert df from get_libor to curve based on month end dates
        #call get_libor, interpolate/extropolate to month_end data points
        #TODO: overload start and end date to allow extrapolation of rates
        df = self.get_libor()
        period = utils.calc_month_end(df['period'].min(), 'date')
        curve = {}
        while period < df.iloc[0]['period']:
            #extropolate backwards - should never happen
            increment = (df.iloc[1]['price'] - df.iloc[0]['price']) / (df.iloc[1]['period'] - df.iloc[0]['period']).days
            interval = (df.iloc[0]['period'] - period).days
            curve[period] = df.iloc[0]['price'] - interval * increment
            period = utils.calc_next_month_end(period, 'date')
        while period <= df['period'].max():
            #interpolate
            bottom_date = max(df.loc[(df['period']<=period)]['period'])
            bottom_yield = df.loc[df['period']==bottom_date]['price'].values[0]
            if period == bottom_date:
                curve[period] = bottom_yield
            elif df.loc[df['period']>period].shape[0] == 0:
                #need to extropolate - does not happen unless overload start and end dates
                increment = (df.iloc[-1]['price'] - df.iloc[-2]['price']) / ((df.iloc[-1]['period'] - df.iloc[-2]['period']).days)
                interval = (period - df.iloc[-1]['period']).days
                curve[period] = df.iloc[-1]['price'] + interval * increment
            else:
                top_date = min(df.loc[(df['period']>=period)]['period'])
                bottom_yield = df.loc[df['period']==bottom_date]['price'].values[0]
                top_yield = df.loc[df['period']==top_date]['price'].values[0]
                increment = (top_yield - bottom_yield) / (top_date - bottom_date).days
                interval = (period - bottom_date).days
                curve[period] = bottom_yield + interval * increment
            period = utils.calc_next_month_end(period, 'date')
            #df_curve = pd.DataFrame(curve, columns= ['period', 'libor'])
        return curve

    def get_amount(self, item, period, cash=0.0, prepay=0.0):
        #TODO cleanup is_interest_payment_date vs is_payment_date
        if item == 'amortization':
            if self.is_interest_payment_date(period) and period >= self.amort_date_start:
                amount =  -self.calc_amort(period)
            else:
                amount = 0.0
        elif item == 'interest expense':
            if self.is_interest_payment_date(period) and period >= self.interest_date_start:
                amount = -self.calc_interest_expense(period, self.principal)
            else:
                amount = 0.0
        elif item == 'sweep':
            amount = 0.0
            if self.flag_prepayable:
                if self.is_interest_payment_date(period):
                    amount = -cash * self.sweep_percent
        elif item == 'interest income':
            if self.is_interest_payment_date(period):
                amount = self.calc_interest_income(period)
            else:
                amount = 0.0
        elif item == 'dsra release':
            #dsra can only be used and therefore reset on a payment date
            if self.is_interest_payment_date(period):
                amount = self.dsra
            else:
                amount =  0.0
        elif item == 'dsra reset':
            #dsra can only be used and therefore reset on a payment date
            if self.is_interest_payment_date(period):
                if self.flag_prepayable:
                    amount = -self.calc_dsra(period, self.principal + prepay)
                else:
                    amount = -self.calc_dsra(period, self.principal)
            else:
                amount = 0.0
        elif item == 'dsra cleanup':
            if self.is_interest_payment_date(period):
                amount = -self.calc_dsra(period, self.principal) + self.dsra
            else:
                amount = 0.0
        else:
            print("Error in FloatingDebt get_amount - unknown item ", item)
            sys.exit()
        return amount

    def calc_interest_rate(self, period):
        try:
            self.interest_rate = self.index[period] + self.margin
        except:
            print("Error in calc_interest_rate - invalid period ", period)
        #if self.label == 'Test TLC':
            #print(self.interest_rate)
        return self.interest_rate


class MezzanineDebt(FixedDebt):
    def __init__(self, **kwargs):
        self.pik_interest_rate = kwargs['pik_interest_rate']
        FixedDebt.__init__(self, **kwargs)

    def get_amount(self, metric, period, cash=0, prepay=0.0):
        #Mezz has cash option (lower interest if paying cash, must pass available cash)
        if metric == 'interest expense':
            #need to determine both cash interest expense and pik interest expense
            #store pik interest to pik list
            if flag_cash_interest == 'standard':
                if self.calc_interest_rate(period) > 0.0:
                    cash_interest = self.calc_interest_expense(period)
                else:
                    cash_interest = 0.0
                if self.pik_interest_rate > 0.0:
                    pik_interest = self.calc_pik_interest(period)
            elif flag_cash_interest == 'optional':
                if cash == 0.0:
                    cash_interest = 0.0
                    pik_interest = self.calc_pik_interest(period)
                else:
                    cash_interest = self.calc_interest_expense(period)
                    pik_interest = max(cash_interest - cash, 0.0) * self.pik_interest_rate /  self.interest_rate
            return cash_interest
        elif metric == 'sweep':
            sweep = cash * self.sweep_percent
            return -sweep
        else:
            print("Error in FloatingDebt get_amount - unknown metric ", metric)
            sys.exit()
        return

    def set_amount(self, item, period, cash_flow):
        if item == 'interest income':
            #not clear if anything needs to happen
            pass
        elif item == 'interest expense':
            self.interest_expense.append((period, cash_flow))
        elif item == 'amortization':
            self.amortization.append((period, cash_flow))
            self.principal += cash_flow
        elif item == 'dsra_change':
            self.dsra_change -= cash_flow
        elif item == 'sweep':
            self.prepayments.append((period, cash_flow))
            self.principal += cash_flow
        elif item == 'pik interest':
            self.principal += cash_flow
        else:
            print("Error - unknow item sent to set_amount ", item)
        return


    def calc_pik_interest(self, period):
        if period > self.maturity_date:
            pik_exp = 0
        elif period < self.payment_date_start:
            pik_exp = 0
        elif self.is_interest_payment_date(period):
            prior_period = self.calc_prior_payment_period(period)
            day_count_factor = utils.calc_day_count_factor(self.day_count, prior_period, period)
            pik_exp = self.principal * self.pik_interest_rate * day_count_factor
        else:
            pik_exp = 0.0

        return int_exp


class Equity:
    def __init__(self, **kwargs):
        self.type = kwargs['class']
        self.label = kwargs['label']
        self.periodicity_months = kwargs['periodicity_months']
        self.first_payment_date = kwargs['first_payment_date'].date()
        self.distributions = {}
        return

    def get_amount(self, item, period, cash=0, prepay=0.0):
        if item == 'sweep':
            if self.is_payment_date(period):
                amount = -cash
            else:
                amount = 0.0
        else:
            print("Error - unknow item sent to Equity get_amount ", item)
        return amount

    def set_amount(self, item, period, cash=0):
        if item == 'sweep':
            self.distributions[period] = cash
        else:
            print("Error - unknow item sent to Equity set_amount ", item)
        return

    def is_payment_date(self,period):
        if period < self.first_payment_date:
            result = False
        elif (period.year * 12 + period.month - self.first_payment_date.year * 12
              - self.first_payment_date.month) % self.periodicity_months == 0:
            result = True
        else:
            result = False
        return result


class FixedAsset:
    def __init__(self, entity, name, tax_life, convention, method, in_service_date, amount, description):
    #def __init__(self, **kwargs):
        self.entity = entity
        self.name = name
        self.tax_life = tax_life
        self.convention = convention
        self.method = method
        self.in_service_date = in_service_date
        self.amount = amount
        self.description = description

    def calc_depreciation(self, period):
        if period < self.in_service_date:
            return 0.0
        elif period > self.in_service_date + relativedelta(years=self.tax_life):
            return 0.0
        else:
            if period.year == self.in_service_date.year:
                if self.convention == 'NA':
                    stub_factor = 0.0
                elif self.convention == 'HY':
                    stub_factor = 0.5
                elif self.convention == 'MM':
                    stub_factor = ((12 - self.in_service_date.month) +.5) / 12
                else:
                    print("ERROR - Unknown convention in calc_depreciation")
                    sys.exit()
            elif period.year == (self.in_service_date + relativedelta(years=self.tax_life)).year:
                if self.convention == 'HY':
                    stub_factor = 0.5
                elif self.convention == 'MM':
                    stub_factor = (period.month - 0.5) / 12
                elif self.convention == 'NA':
                    stub_factor = 0.0
                else:
                    print("ERROR - Unknown convention in calc_depreciation")
                    sys.exit()
            else:
                stub_factor = 1.0

            if self.method == 'SL':
                method_factor = 1 / self.tax_life
            elif self.method == 'MACRS':
                #add later
                pass
            elif self.method == 'NA':
                method_factor = 0.0
            else:
                print("ERROR - Unknown method in calc_depreciation", method)
                sys.exit()

        return stub_factor * method_factor * self.amount


class TaxRegister:
    def __init__(self, name):
        self.entity = name
        self.assets = []

    def add_asset(self, asset):
        self.assets.append(asset)
        return

    def calc_depreciation(self, period):
        #returns annual tax depreciation for the period
        depreciation = 0.0
        for asset in self.assets:
            depreciation += asset.calc_depreciation(period)
        return depreciation

    def print_assets(self):
        for asset in self.assets:
            print(asset.entity, asset.name, asset.tax_life, asset.convention, asset.method, asset.in_service_date, asset.amount, asset.description)
        return


def get_portfolio_from_xlsx():
    #template is hard coded here
    path_data = str(Path(path).parent) + '/data/lbo/'
    wb = opx.load_workbook(path_data + 'assumptions_template.xlsx')
    ws = wb['portfolio']
    portfolio_kwargs = {}
    row = 5
    while ws['a'+str(row)].value is not None:
        key = ws['a' + str(row)].value
        value = ws['b' + str(row)].value
        portfolio_kwargs[key] = value
        row += 1
    wb.close()
    return portfolio_kwargs

def get_cap_struct_from_xlsx():
    #template is hard coded here
    path_data = str(Path(path).parent) + '/data/lbo/'
    wb = opx.load_workbook(path_data + 'assumptions_template.xlsx')
    ws = wb['capital structure']
    #scenario/version fixed location
    cap_struct_scenario = ws['b3'].value
    cap_struct_version = ws['b4'].value
    cap_struct = {}
    #name always starts at row 6, variable numbers of kwargs
    instrument_key = ws['b6'].value
    instrument = {}
    row = 7
    while ws['a'+str(row)].value is not None:
        key = ws['a' + str(row)].value
        if key == 'name':
            #close out prior dictionary item
            cap_struct[instrument_key] = instrument
            instrument_key = ws['b' + str(row)].value
            instrument = {}
        else:
            instrument[key] =  ws['b' + str(row)].value
        row += 1
    #final entry in dictionary
    cap_struct[instrument_key] = instrument
    wb.close()
    return cap_struct

def get_waterfall_from_xlsx():
    #template is hard coded here
    path_data = str(Path(path).parent) + '/data/lbo/'
    wb = opx.load_workbook(path_data + 'assumptions_template.xlsx')
    ws = wb['waterfall']

    flow = namedtuple('flow', ('level, sublevel, instrument, item, method, split, flag_cash, report_subtotal'))
    waterfall = []
    row = 2
    while ws['a'+str(row)].value is not None:
        #scenario = ws['a' + str(row)].value
        #version = ws['b' + str(row)].value
        level = ws['c' + str(row)].value
        sublevel = ws['d' + str(row)].value
        instrument = ws['e' + str(row)].value
        item = ws['f' + str(row)].value
        method = ws['g' + str(row)].value
        split = ws['h' + str(row)].value
        flag_cash = ws['i' + str(row)].value
        report_subtotal = ws['j' + str(row)].value
        waterfall.append(flow(level, sublevel, instrument, item, method, split, flag_cash, report_subtotal))
        row += 1
    wb.close()
    return waterfall

def get_tax_register_from_xlsx():
    #template is hard coded here
    #returns a list of tax assets
    path_data = str(Path(path).parent) + '/data/lbo/'
    wb = opx.load_workbook(path_data + 'assumptions_template.xlsx')
    ws = wb['taxes']

    fixed_asset = namedtuple('fixed_asset', ('entity, name, tax_life, convention, method, in_service_date, amount, description'))
    register = []
    row = 2
    while ws['a'+str(row)].value is not None:
        entity = ws['a' + str(row)].value
        name = ws['b' + str(row)].value
        tax_life = ws['c' + str(row)].value
        convention = ws['d' + str(row)].value
        method = ws['e' + str(row)].value
        in_service_date = ws['f' + str(row)].value.date()
        amount = ws['g' + str(row)].value / UNITS
        description = ws['h' + str(row)].value
        #register.append(fixed_asset(entity, name, tax_life, convention, method, in_service_date, amount, description))
        register.append([entity, name, tax_life, convention, method, in_service_date, amount, description])
        row += 1
    wb.close()
    return register

def npv(irr, cfs, yrs):
    return np.sum(cfs / (1. + irr)**yrs)

def irr(cfs, yrs, x0, **kwargs):
    return np.asscalar(fsolve(npv, x0=x0, args=(cfs, yrs), **kwargs))

def required_return_payment(payment, payment_years, investment, irr, cfs, yrs):
    return npv(irr, cfs, yrs) - investment + npv(irr, payment, payment_years)

def convert_to_years(dividends, investment_date):
    years = []
    for dividend in dividends:
        years.append((dividend[1]-investment_date)/365)
    return years

def load_instruments(scenario):
    #returns a dictionary of debt/equity instruments (objects)
    instruments = {}
    for ins in scenario:
        kwargs = scenario[ins]
        #print(kwargs['flag_include'])
        #sys.exit()
        if kwargs['flag_include']:
            instruments[ins] = globals()[kwargs['class']](**kwargs)
    return instruments

def pari_passu(period, level, sublevel, cash, waterfall, instruments, output):
    #identify sublist from portfolio.waterfall of items that are pari passu
    pari_items = []
    for flow in waterfall:
        if flow.level == level and flow.sublevel == sublevel:
            pari_items.append(flow)
    #cycle thru sublist with get_amount to determine total request
    cash_requested = 0.0
    for flow in pari_items:
        cash_requested += instruments[flow.instrument].get_amount(flow.item, period, cash)
    #calc pro-rata amount
    if cash_requested == 0.0:
        pro_ration = 1.0
    else:
        pro_ration = max(cash / cash_requested, 1.0)
    #cycle thru sublist with set_amount
    for flow in pari_items:
        cash_flow = instruments[flow.instrument].get_amount(flow.item, period, cash)
        if flag_debug:
            print("{:,.2f}".format(cash), flow.instrument, flow.item, "{:,.2f}".format(cash_flow), 'pari passu')
        portfolio.instruments[flow.instrument].set_amount(flow.item, period, cash_flow * pro_ration)
        cash += cash_flow * pro_ration
        output.append([period, flow.instrument, flow.item, cash, cash_flow * pro_ration, flow.level, flow.sublevel])
    #return remaining cash
    return cash, output

def calc_next_flow(level, sublevel, waterfall):
    next_level = 0
    next_sublevel = 0
    flag = False
    for flow in waterfall:
        if flow.level == level and flow.sublevel == sublevel:
            flag = True
        else:
            if flag == True:
                next_level = flow.level
                next_sublevel = flow.sublevel
                break
    return (next_level, next_sublevel)

def run_waterfall(prepay_solver, portfolio):
    #This is run to get the initial estimate of prepayments to feed the solver
    period = utils.calc_month_end(portfolio.close_date, 'date')
    month = 0
    periodicity_months = 1
    output = []
    excess_cash = []
    while period <= utils.calc_month_end(portfolio.terminal_date, 'date'):
        cash = 0.0
        if flag_debug:
            print(period)
        #first flow is always level 1, sublevel 1
        next_flow = (1,1)
        for flow_counter in range(len(portfolio.waterfall)):
            #note: cash outflows are negative, inflows positive
            flow = portfolio.waterfall[flow_counter]
            if (flow.level * 100 + flow.sublevel) == (next_flow[0] * 100 + next_flow[1]):
                if flow.method == 'normal':
                    #special case if calling portfolio-level function (necessary for ptd)
                    if flow.instrument == 'Portfolio':
                        cash_requested = portfolio.get_amount(flow.item, period, cash, prepay_solver)
                    else:
                        cash_requested = portfolio.instruments[flow.instrument].get_amount(flow.item, period, cash, prepay_solver[month])
                    if cash_requested >= 0:
                        cash_flow = cash_requested
                    else:
                        #cleanup items need to allow for negative cash flow.  Will be zero when solution found
                        if flow.item in ['dsra cleanup', 'ptd cleanup']:
                            cash_flow = cash_requested
                        else:
                            cash_flow = max(cash_requested, -cash)
                    if flow.instrument == 'Portfolio':
                        portfolio.set_amount(flow.item, period, cash_flow)
                    else:
                        portfolio.instruments[flow.instrument].set_amount(flow.item, period, cash_flow)
                    output.append([period, flow.instrument, flow.item, cash, cash_flow, flow.level, flow.sublevel])
                    if flag_debug:
                        print("{:,.2f}".format(cash), flow.instrument, flow.item, "{:,.2f}".format(cash_flow), "{:,.2f}".format(portfolio.instruments['TLB'].principal))
                    if flow.flag_cash:
                        cash += cash_flow
                elif flow.method == 'pari passu':
                    cash, output = pari_passu(period, flow.level, flow.sublevel, cash, portfolio.waterfall, portfolio.instruments, output)
                else:
                    print('Error in main loop - unknown waterfall item')
            next_flow = calc_next_flow(flow.level, flow.sublevel, portfolio.waterfall)
        period = utils.calc_next_month_end(period, 'date')
        month += 1
    return output

def solve_waterfall(prepay_solver, portfolio):
    #This is run to get the initial estimate of prepayments to feed the solver
    period = utils.calc_month_end(portfolio.close_date, 'date')
    month = 0
    periodicity_months = 1
    output = []
    excess_cash = []
    while period <= utils.calc_month_end(portfolio.terminal_date, 'date'):
        cash = 0.0
        if flag_debug:
            print(period)
        #first flow is always level 1, sublevel 1
        next_flow = (1,1)
        for flow_counter in range(len(portfolio.waterfall)):
            #note: cash outflows are negative, inflows positive
            flow = portfolio.waterfall[flow_counter]
            if (flow.level * 100 + flow.sublevel) == (next_flow[0] * 100 + next_flow[1]):
                if flow.method == 'normal':
                    #special case if calling portfolio-level function (necessary for ptd)
                    if flow.instrument == 'Portfolio':
                        cash_requested = portfolio.get_amount(flow.item, period, cash, prepay_solver)
                    else:
                        cash_requested = portfolio.instruments[flow.instrument].get_amount(flow.item, period, cash, prepay_solver[month])
                    if cash_requested >= 0:
                        cash_flow = cash_requested
                    else:
                        #cleanup items need to allow for negative cash flow.  Will be zero when solution found
                        if flow.item in ['dsra cleanup', 'ptd cleanup']:
                            cash_flow = cash_requested
                        else:
                            cash_flow = max(cash_requested, -cash)
                    if flow.instrument == 'Portfolio':
                        portfolio.set_amount(flow.item, period, cash_flow)
                    else:
                        portfolio.instruments[flow.instrument].set_amount(flow.item, period, cash_flow)
                    output.append([period, flow.instrument, flow.item, cash, cash_flow, flow.level, flow.sublevel])
                    if flag_debug:
                        print("{:,.2f}".format(cash), flow.instrument, flow.item, "{:,.2f}".format(cash_flow), "{:,.2f}".format(portfolio.instruments['TLB'].principal))
                    if flow.flag_cash:
                        cash += cash_flow
                elif flow.method == 'pari passu':
                    cash, output = pari_passu(period, flow.level, flow.sublevel, cash, portfolio.waterfall, portfolio.instruments, output)
                else:
                    print('Error in main loop - unknown waterfall item')
            next_flow = calc_next_flow(flow.level, flow.sublevel, portfolio.waterfall)
        excess_cash.append(cash)
        period = utils.calc_next_month_end(period, 'date')
        month += 1
    return excess_cash

def create_portfolio():
    portfolio_kwargs = get_portfolio_from_xlsx()
    portfolio = Portfolio(**portfolio_kwargs)

    #cap_struct is a dictionary of dictionaries
    #   first key is name of instrument
    #   second key is kwarg for object
    portfolio.cap_struct = get_cap_struct_from_xlsx()

    #instruments is a dictionary of objects
    #   key is name of instrument
    portfolio.instruments = load_instruments(portfolio.cap_struct)


    #waterfall is a list of namedtuples
    #   each item in the list is a step in the waterfall
    portfolio.waterfall = get_waterfall_from_xlsx()
    return portfolio

def waterfall_shell(prepay_solver):
    global COUNTER
    print(COUNTER)
    COUNTER += 1
    #prepay_solver is a list of prepayments
    portfolio = create_portfolio()
    #if portfolio.flag_ptd:
    #    portfolio.calc_ptd()
    excess = solve_waterfall(prepay_solver, portfolio)
    time_elapesed = datetime.now() - START_TIME
    #print(time_elapesed)
    return excess

if __name__ == '__main__':
    #TODO: figure out global database access
    HOST = 'kindledb.cfdmlfy5ocmf.us-west-2.rds.amazonaws.com'
    USER = 'Andrew'
    PASSWORD = 'Kindle01'
    DATABASE = 'kean'
    cnx = utils.generate_connection_instance(HOST, USER, PASSWORD, DATABASE)

    global START_TIME
    START_TIME = datetime.now()

    global UNITS
    UNITS = 1000000

    flag_debug = True

    #need global variable for scenario start and end dates
    #try:
    #    PORTFOLIO_START_DATE = portfolio.close_date
    #except:
    #    print("Error - no portfolio start date selected")

    #need global variable for yield curve date.  attribute of Portfolio
    try:
        YIELD_CURVE_DATE = date(2019, 6, 26)    #portfolio.yield_curve_date
        YIELD_CURVE_VERSION = 'v3'              #portfolio.yield_curve_version
    except:
        #not using KEAN for libor curve
        pass


    #create prepay_solver
    prepay_solver = []
    period = date(2019, 6, 30)
    terminal_date = date(2020, 12, 31)

    #establish initial estimate of prepayments
    #   can change value in 'if' portion of statement
    while period <= terminal_date:
        if period == date(2019, 6, 30):
            prepay_solver.append(0.0)
        else:
            prepay_solver.append(0.0)
        period = utils.calc_next_month_end(period, 'date')

    #below solver was attempt to speed up fsolve by using solution as first estimate, did not work
    #prepay_solver = [0,0,0,-30.2118,0,0,-19.4408,0,0,-69.8305,0,0,0,0,0,-24.4698,0,0,-10.0611]
    portfolio = create_portfolio()
    output = run_waterfall(prepay_solver, portfolio)
    #print('amortization ', portfolio.instruments['TLC'].amortization)
    #print('prepayments ', portfolio.instruments['TLC'].prepayments)
    #print('principal', portfolio.instruments['TLC'].principal, portfolio.instruments['TLC'].initial_balance)
    #sys.exit()

    df_output = pd.DataFrame(output, columns=['period', 'instrument', 'item', 'cash', 'cash_flow', 'level', 'sublevel'])
    criteria = ((df_output['item']=='sweep') & (df_output['instrument']=='TLB'))
    prepay_solver = df_output[criteria]['cash_flow'].tolist()

    #THIS IS WHERE THE MAGIC HAPPENS
    COUNTER = 1
    portfolio = create_portfolio()
    solver = fsolve(waterfall_shell, prepay_solver)

    #save solver solution to csv file to use later in testing
    #with open('solver.csv', 'w', newline='') as myfile:
    #    wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
    #    wr.writerow(solver)

    portfolio = create_portfolio()
    output = run_waterfall(solver, portfolio)
    df_output = pd.DataFrame(output, columns=['period', 'instrument', 'item', 'cash', 'cash_flow', 'level', 'sublevel'])
    #df_output.to_csv('lbo_output.csv')
    #create_lbo_support_report(portfolio)
    create_waterfall_report(df_output, portfolio.waterfall)
